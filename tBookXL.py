"""---------------------------------------------------------------------------------------------------------------------
---------------------------------------------[ TableBookXL ]------------------------------------------------------------
---------------------------------------------------------------------------------------------------------------------"""

import os
import xlrd
from openpyxl import Workbook, load_workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.utils import get_column_letter
from alx import add_str, class_name, quote, inspect_info, dbg, get_datetime_stamp, del_if_exist, is_opened

from tLib import xls_ext, xlsx_ext, TableError, read_xlsx_sheet, read_xls_sheet_data, reorder_dict, get_test_indexes
from tFormat import Formatter
from tTable import Table
from tBook import TableBook




class TableBookXL:  # TableBookXL

    index = 0

    def __init__(self, source=TableBook()):  # TableBookXL
        """
        Parameters
        ----------
        table_book : объект класса TableBook
        file : если объект TableBookXL создается для существующего файла - необходимо указать имя файла.
        """

        """ ws_dic имеет следующий формат после __init__(): {title: openpyxl.worksheet,...,
                                                             titleN: openpyxl.worksheetN}
            ft_dic имеет следующий формат после __init__(): {}

            Если имя файла не было указано, то используется table_book для создания пустых листов excel. Заполнение 
            листов данными из table_book происходит только при сохранении файла excel. Также, при сохранении может быть 
            использован форматтер, для каждого листа свой, если он был указан перед сохранением вызовом метода 
            format_sheet(<title>).

            Если имя файла было указано, то создается новый объект table_book, который заполняется информацией из файла
            Excel.    

            Если имя файла указано, table_book заполняется информацией из файла.    
        """

        self.index = 0

        def __file_mode(flag=None):  # TableBookXL
            if flag is not None:
                self.src_file_mode = flag
            return self.src_file_mode

        def __book_mode(flag=None):  # TableBookXL
            if flag is not None:
                self.src_file_mode = not flag
            return not self.src_file_mode

        self.src_file_mode = None  # True = got from file
        self.src_file_xls = None  # True  - .xls формат, use XLRD else OpenPyXL

        self.src_wb = None
        self.src_ws_dic = {}
        self.src_filename = None

        self.dst_wb = None
        self.dst_ws_dic = {}
        self.dst_filename = None

        self.ft_dic = {}

        if isinstance(source, TableBook):
            __book_mode(True)
            self.table_book: TableBook = source
            self.src_filename = None
        elif isinstance(source, str):
            __file_mode(True)
            self.table_book = TableBook()
            self.src_filename = os.path.abspath(source)
            src_file_ext = os.path.splitext(self.src_filename)[1].lower()

            if src_file_ext == xls_ext.lower():
                self.src_file_xls = True
            elif src_file_ext == xlsx_ext.lower():
                self.src_file_xls = False
            else:
                raise TableError(f'{class_name(self)}: Файл должен иметь расширение {xls_ext} или {xlsx_ext}. Получен: '
                                 f'{self.src_filename}\n{inspect_info()}')

            if not os.path.isfile(self.src_filename):
                raise TableError(f'{class_name(self)}: Файл {self.src_filename} не найден.\n'
                                 f'{inspect_info()}')
        else:
            raise TableError(f'{class_name(self)}: Получен входной объект, который не является объектом класса'
                             f'<TableBook> или <именем файла:str>.\n'
                             f'{inspect_info()}')

        """ Создаем workbook openpyxl в памяти для сохранения в файл, если это будет далее необходимо """

        """ создаем workbook excel в памяти """
        self.dst_wb = Workbook(write_only=True)

        if __book_mode():
            """ На входе у нас TableBook - мы знаем имена листов выходного xls файла и можем их создать,
                пока пустыми """
            for title, table in self.table_book:
                self.dst_ws_dic[title] = None  # пока только резервируем title в словаре. Листы openpyxl создадим
                # перед сохранением, потому что нас предусмотрен метод удаления листов.
        else:
            """ На входе у нас xls или xlsx файл. Считываем данные листов, создаем объекты Table, заносим их 
                в TableBook """

            """ если .xls - используем xlrd """
            if self.src_file_xls:
                self.src_wb = xlrd.open_workbook(self.src_filename)

                ws_number = self.src_wb.nsheets

                for ws_index in range(ws_number):
                    ws = self.src_wb.sheet_by_index(ws_index)
                    title = ws.name
                    table_data = read_xls_sheet_data(ws)
                    self.src_ws_dic[title] = ws
                    self.table_book.new_table(title, table_data)

                """ считываем данные листов, создаем объекты Table, заносим их в TableBook """
                """ если .xlsx - используем openpyxl """
            else:
                self.src_wb = load_workbook(self.src_filename, read_only=True, data_only=True)

                for ws in self.src_wb:
                    title = ws.title
                    table_data = read_xlsx_sheet(ws)
                    self.src_ws_dic[title] = ws
                    self.table_book.new_table(title, table_data)

            """ Теперь, когда имена листов известны и сохранены в TableBook - создаем соответсвующее листы выходного 
                файла. И заполняем словарь форматтера"""

            for title in self.table_book.list_titles():
                self.dst_ws_dic[title] = self.dst_wb.create_sheet(title)
                self.ft_dic[title] = Formatter()

    def __contains__(self, title):  # TableBookXL
        return title in self.table_book.tables_dic

    def __getitem__(self, indx: int | str) -> tuple[str | int, Table, Formatter]:  # TableBookXL
        """
        Parameters
        ----------
        indx : если i - целое число, то метод возвращает имя таблицы в книге и саму таблицу
            если i - имя таблицы в книге, то метод возвращает индекс таблицы и саму таблицу
        Returns
        -------

        """
        _titles = self.table_book.list_titles()

        if type(indx) == str:
            if indx in _titles:
                _title, _table = self.table_book[indx]
                return _title, _table, self.ft_dic[indx]
            else:
                raise TableError(f'{class_name(self)}: Таблица "{indx}" не найдена.\n'
                                 f'{inspect_info()}')
        elif type(indx) == int:
            if -1 < indx < len(_titles):
                _title, _table = self.table_book[indx]
                return _title, _table, self.ft_dic[_titles[indx]]
            else:
                raise TableError(f'{class_name(self)}: Таблица с индексом [{indx}] не найдена.\n'
                                 f'{inspect_info()}')

    def __len__(self) -> int:  # TableBookXL
        return len(self.table_book)

    def __iter__(self):  # TableBookXL
        self.index = 0
        return self

    def __next__(self):  # TableBookXL
        if self.index >= self.__len__():
            raise StopIteration()
        else:
            self.index += 1
            return self.__getitem__(self.index - 1)

    def _check_title_(self, title='', inclusion=True) -> str:  # TableBookXL
        title = str(title).strip()
        _titles = self.table_book.list_titles()

        if title == '':
            if len(_titles) > 0:
                title = _titles[0]

        if inclusion:
            if title not in self.table_book:
                raise TableError(f'{class_name()}: Лист или таблица с именем {title} не найдены.\n'
                                 f'{inspect_info()}')
        else:
            if title in self.table_book:
                raise TableError(f'{class_name()}: Лист или таблица с именем {title} уже существуют.\n'
                                 f'{inspect_info()}')
        return title

    def auto_format(self, title):
        """
        Настраивает тип столбцов по данным столбцов (по выборке до 9 строк из таблицы).
        Parameters
        ----------
        title :

        Returns
        -------

        """

        def test_row(data, row_index, formatter):
            key_list = []
            title_list = []
            type_list = []
            if type(row_index) == int:
                if -1 < row_index < len(data):
                    pass
                else:
                    raise TableError(f'{class_name(self)}: Параметр row_index указывает за пределы таблицы данных'
                                     f'{inspect_info()}')
            else:
                raise TableError(f'{class_name(self)}: Параметр row_index должен иметь тип int.'
                                 f'{inspect_info()}')
            if isinstance(data[row_index], list) or isinstance(data[row_index], tuple):
                for col_index in range(len(data[row_index])):  # определяем число столбцов по строке data[row_index]
                    col_key = get_column_letter(col_index + 1)
                    key_list.append(col_key)
                    title_list.append(col_index + 1)
            elif isinstance(data[row_index], dict):
                for col_index, col_key in enumerate(data[row_index].keys()):
                    key_list.append(col_key)
                    title_list.append(col_index + 1)
            else:
                raise TableError(f'{class_name(self)}: Строка data[row_index] не содержит строк в формате '
                                 f'списков, кортежей или словарей.\n'
                                 f'{inspect_info()}')
            """ Ключи и заголовки готовы. Теперь определим типы """
            type_list = formatter.guess_col_types(data[row_index], keys_ready=False)  # keys_ready = False, потому что
            return key_list, title_list, type_list

        ws_formatter = self.get_formatter(title)
        ws_table = self.table_book.get_table(title)
        test_row_indexes = get_test_indexes(ws_table.rows)

        # TO BE CONTINUED ...

        # for i in test_row_indexes:
        #     keys, titles, types = test_row(array, i)

    def get_formatter(self, title='') -> Formatter:  # TableBookXL
        """
        Возвращает фактическую ссылку на форматтер (поскольку он mutable), для обращения к нему его методами.
        Parameters
        ----------
        title : имя листа

        Returns
        -------
        Formatter
        """
        title = self._check_title_(title)
        return self.ft_dic[title]

    def tablebook(self):
        return self.table_book

    def _rearrange_(self, titles_list: list[str]) -> None:
        titles_list = list(self._check_title_(title) for title in titles_list)
        self.table_book.reorder_titles(titles_list)
        self.ft_dic = reorder_dict(self.ft_dic, titles_list)
        self.dst_ws_dic = reorder_dict(self.dst_ws_dic, titles_list)

    def add_sheet(self, title: str, table: Table, formatter=Formatter()) -> None:
        title = self._check_title_(title, inclusion=False)

        self.table_book.add_table(title, table)
        self.dst_ws_dic[title] = None  # создаем листы openpyxl только перед сохранением, потому что мы можем их удалять
        self.ft_dic[title] = formatter

        self._rearrange_(self.table_book.list_titles())

    def create_sheet(self, title: str, table_data, keys=0, columns=None, titles=None, filler=None,
                     formatter=Formatter()) -> None:
        title = self._check_title_(title, inclusion=False)

        self.table_book.new_table(title, table_data, keys, columns, titles, filler)
        self.dst_ws_dic[title] = None  # создаем листы openpyxl только перед сохранением, потому что мы можем их удалять
        self.ft_dic[title] = formatter

        self._rearrange_(self.table_book.list_titles())

    def pop_sheet_tuple(self, title) -> tuple[str, Table, Formatter]:

        title = self._check_title_(title)
        table = self.table_book.get_table(title)
        formatter = self.ft_dic[title]

        self.table_book.del_table(title)
        self.ft_dic.pop(title)
        self.dst_ws_dic.pop(title)

        self._rearrange_(self.table_book.list_titles())
        return title, table, formatter

    def set_formatter(self, title='', formatter=None) -> None:
        title = self._check_title_(title)
        if formatter is not None:
            self.ft_dic[title] = formatter

    def fill_xl_sheet(self, title: str, header=False) -> None:
        """
        Заполняет и форматирует openpyxl лист Excel
        Parameters
        ----------
        title : название листа
        header : Если True, то в лист добавляется заголовок таблицы.

        Returns
        -------
        None
        """

        def get_styled_row(ws_obj, row_data_list, row_data_index):
            """
            Устанавливает стиль для строки. Не трогает типы столбцов, поскольку они будут назначены поколоночно,
            когда лист будет сформирован (так работает openpyxl - тип столбца для всего столбца может быть задан
            только когда столбец полностью сформирован. Альтернатива - назначить тип каждой ячейки, но это,
            предположительно, более ресурсоемкая операция.

            Parameters
            ----------
            ws_obj: openpyxl Worksheet
            row_data_list : строка данных
            row_data_index :

            Returns
            -------

            """
            styled_row = []
            for col_index in range(len(row_data_list)):
                cell = WriteOnlyCell(ws_obj, value=row_data_list[col_index])
                cell.style = formatter.cstyle(row_data_index, col_index)
                # cell.number_format = formatter.ctype(col_index) (назначение типов перенесено в set_columns_types())
                styled_row.append(cell)
            return styled_row

        def set_columns_types(ws_obj, ws_table):
            for ws_col_index in range(ws_table.columns):
                ws_column = ws_obj.column_dimensions[get_column_letter(ws_col_index + 1)]
                ws_column.number_format = formatter.ctype(ws_col_index)

        title = self._check_title_(title)
        ws = self.dst_ws_dic[title]
        idx, table_obj = self.table_book[title]
        formatter = self.ft_dic[title]

        if header:
            for row_index in range(len(table_obj) + 1):
                if row_index == 0:
                    row_lst = table_obj.get_header()
                    ws.append(get_styled_row(ws, row_lst, 0))
                else:
                    row_lst = table_obj[row_index - 1].get_row_data()
                    ws.append(get_styled_row(ws, row_lst, row_index))
        else:
            for row_index in range(len(table_obj)):
                row_lst = table_obj[row_index].get_row_data()
                ws.append(get_styled_row(ws, row_lst, row_index))

        """ Теперь, когда лист заполнен данными и стилями, установим типы столбцов. Предположительно, это будет быстрее,
            чем устанавливать тип каждой ячейки отдельно."""
        set_columns_types(ws, table_obj)

    def save(self, filename: str, header=False):
        if not is_opened(filename):
            for title, table in self.table_book:
                self.dst_ws_dic[title] = self.dst_wb.create_sheet(title)
                self.fill_xl_sheet(title, header=header)
            self.dst_wb.save(filename)
        else:
            raise TableError(f'{class_name(self)}: Файл {filename} открыт в другом приложении и не может быть '
                             f'перезаписан.\n{inspect_info()}')
        self.close()

    def close(self):
        self.dst_wb.close()
        if self.src_file_xls is not None:
            if self.src_file_xls:
                self.src_wb.release_resources()
            else:
                self.src_wb.close()

