"""---------------------------------------------------------------------------------------------------------------------
---------------------------------------------[ class Formatter ]--------------------------------------------------------
---------------------------------------------------------------------------------------------------------------------"""
from copy import deepcopy as duplicate
import openpyxl.worksheet.worksheet
from openpyxl.cell import WriteOnlyCell
from openpyxl.utils import get_column_letter

from alx import add_str, class_name, quote, inspect_info
from tLib import TableError, xl_cell_type_gen, xl_cell_type_int, xl_cell_type_text, xl_cell_type_float, \
    xl_cell_style_simple, excel_default_style, excel_default_type

ft_col_formatter = 'col'
ft_row_formatter = 'row'
formatter_template = {ft_col_formatter: dict(), ft_row_formatter: dict()}
ft_i_col_title = 0  # [,column title,,]
ft_i_col_type = 1  # [,,column type format (format_text or format_percent or format_float or format_general)]
ft_i_col_value = 2  # [,,,anything you want to store]
ft_i_col_func = 3  # [,,,function to apply to value, for example to convert types]
ft_col_title = 'title'
ft_col_type = 'type'
ft_col_default = 'default'
ft_col_func = 'func'
ft_row_style = 'style'
ft_row_font = None  # at the moment
ft_i_row_style = 0


class Formatter:  # Formatter
    """
    Принцип форматтера состоит в том, что он задает формат данных (тип данных) поколоночно,
    а формат ячеек (стиль ячеек в терминах excell) - построчно,
    Т.е.:
        формат данных - это форматы данных по типу данных (текст, целое, плавающее, проценты, и пр.
                        (см. https://openpyxl.readthedocs.io/en/stable/_modules/openpyxl/styles/numbers.html),

        формат ячеек -  это форматы ячеек (цвет, бордюр) в соответствии со встроенными стилями
                        (см. excel - styles->cell styles +
                        https://openpyxl.readthedocs.io/en/stable/styles.html )
    ---------------------------------------------------------------------------------------------------------
    Форматтер состоит из

        - словаря форматов столбцов
                {
                  <ключ столбца 0>: {<атрибут 1>:<значение 1>, ..., <атрибут n>:<значение n>},
                  ...
                  <ключ столбца N>: {<атрибут 1>:<значение 1>, ..., <атрибут n>:<значение n>}
                  }

                  Перечень аттрибутов столбца: ft_col_title, ft_col_type, ft_col_default, ft_col_func
                  при инициализации объектов присваиваются значения ключей столбцов и значение заголовка столбца
                  (аттрибут ft_col_title)

        - словаря форматов (бордюр, цвет) сегментов строк.
           {
             (<Номер строки>, <ключ столбца начала сегмента>): {<атрибут 1>:<значение 1>,..,<атрибут n>:<значение 1>},
              ....
             (<Номер строкиN>, <ключ столбца начала сегмента>): {<атрибут N>:<значение N>,..,<атрибут n>:<значение N>},

            Т.е., ключ строки - это кортеж из (<номера строки или -1>, <ключ столбца, с которого начинается сегмент>)
               Номер строки = -1 - применяется для форматирования тех строк, о которых нет своей записи в словаре,
               т.е. для всех остальных строк.

            Перечень аттрибутов строки включает:
                ft_row_style:<один из встроенных стилей excell> для ячейки).
                ft_row_font: 0 # сейчас не используется. Добавлен для целей отладки

            * При инициализации объектов:
                    В словаре столбцов: присваиваются значения ключей столбцов и значение заголовка столбцов
                                        (аттрибут ft_col_title)
                    В словаре строк:    Первой строке (номер строки = 0) применяется стиль xl_cell_style_title =
                                        'Check Cell',
                                        Остальным строкам применяется стиль xl_cell_style_simple = 'Output'.
    """

    template_was_used = False

    col_formatter: dict = {}
    row_formatter: dict = {}

    columns: int = 0  # устанавливается при инициализации и не изменяется
    rows: int = 0  # Хранит количество строк последней отформатированной таблицы или 0

    ix: dict = {}
    cached_types_list: list = []
    cached_styles_dic: dict = {}

    """---------------------------------------------------------------------------------------------------------------- 
    *** __*__ METHODS >>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>> 
    ----------------------------------------------------------------------------------------------------------------"""

    def __init__(self, template_data=None, template_row=0, columns=0, **keys_titles):  # Formatter
        """
        Предусмотрено четыре взаимоисключающих способа передачи параметров:
        0) Без параметров: создается пустой форматтер.
        1) key_titles = {ключ1:заголовок1,...,ключN:заголовокN}
        2) template_table = ссылка на таблицу (list), которая может иметь форматы: список списков или список словарей.
              * Если использован список списков - то форматтер настраивается на число столбцов по template_row строке
              таблицы, а ключи столбцов генерируются по правилам Excel - 'A', "B', и т.д. Заголовки содержат номера
              столбцов от 1.
              * Если использован список словарей - то форматтер настраивается по первому словарю. Ключи словаря
                используются как ключи строк. Заголовки содержат номера столбцов от 1.
        3) columns = <число столбцов>) - форматтер настраивается на число столбцов, а ключи столбцов генерируются по
            правилам Excel - 'A', "B', и т.д. Заголовки содержат номера столбцов, начиная с 1 (а не 0!)

        Parameters
        ----------
        template_data :  Образец таблицы, на которую настраивается форматтер. Настройка идет по первой строке.
        columns : Число столбцов, на который настраивается форматтер.
        keys_titles : Ключевые параметры, задающие ключ и значение заголовка каждого столбца в формате ключ1=заголовок1,...
                  ...,ключN=заголовокN
        """
        self.template_was_used = False

        self.col_formatter: dict = {}
        self.row_formatter: dict = {}

        self.columns: int = 0  # устанавливается при инициализации и не изменяется
        self.rows: int = 0  # Хранит количество строк последней отформатированной таблицы или 0

        self.ix: dict = {}
        self.cached_types_list: list = []
        self.cached_styles_dic: dict = {}

        if template_data and not (columns > 0 or keys_titles):  # template не пустой

            self.template_was_used = True

            if type(template_row) == int:
                if -1 < template_row < len(template_data):
                    pass
                else:
                    raise TableError(f'{class_name(self)}: Параметр template_row указывает за пределы таблицы'
                                     f'{inspect_info()}')
            else:
                raise TableError(f'{class_name(self)}: Параметр template_row должен иметь тип int.'
                                 f'{inspect_info()}')

            if isinstance(template_data[template_row], list) or isinstance(template_data[template_row], tuple):
                for col in range(len(template_data[template_row])):  # определяем число столбцов по строке template_row
                    col_key = get_column_letter(col + 1)
                    keys_titles.update({col_key: col + 1})
            elif isinstance(template_data[template_row], dict):
                for col_index, col_key in enumerate(template_data[template_row].keys()):
                    keys_titles.update({col_key: col_index + 1})
            else:
                raise TableError(f'{class_name(self)}: Строка template_data[template_row] не содержит строк в формате '
                                 f'списков, кортежей или словарей.\n'
                                 f'{inspect_info()}')

        elif (columns > 0) and not (template_data or keys_titles):
            for col in range(columns):  # размер (число столбцов) определяем по первой строке
                col_key = get_column_letter(col + 1)
                keys_titles.update({col_key: col + 1})

        elif keys_titles and not (columns > 0 or template_data):
            pass
        elif not (keys_titles or columns > 0 or template_data):
            pass
        else:
            raise TableError(f'{class_name(self)}: Конфликт параметров. Допускается указание или шаблона template_data'
                             f' или числа столбцов columns или перечня ключей со значениями заголовков.\n'
                             f'{inspect_info()}')

        self.col_formatter: dict = dict()
        self.row_formatter: dict = dict()

        self.columns: int = len(keys_titles)
        self.rows: int = 0

        """ инициализируем кэш, который состоит из:
            - словаря индексов: ix  (трансляция ключей столбцов в номера столбцов и наоборот)
            - списка типов столбцов: col_types  
            - словаря стилей сегментов строк: row_styles
        """
        self.ix: dict = dict()
        self.cached_types_list: list = list()
        self.cached_styles_dic: dict = dict()

        if len(keys_titles) > 0:
            for key, value in keys_titles.items():
                self.col_formatter.update(
                    {
                        key:
                            {
                                ft_col_title: value,
                                ft_col_type: xl_cell_type_gen,
                                ft_col_default: "",
                                ft_col_func: None
                            }
                    })

            """ теперь зададим формат (стиль) строк по умолчанию """
            self.sync_indexes()  # !

            self.row_formatter.update({(-1, self.ix[0]): {}})
            self.row_formatter[(-1, self.ix[0])].update({ft_row_style: xl_cell_style_simple})
            self.row_formatter[(-1, self.ix[0])].update({ft_row_font: None})

        else:  # если столбцов нет, то ключ столбца форматтера строки = None
            self.ix = {-1: None, None: -1}

            self.row_formatter.update({(-1, None): {}})
            self.row_formatter[(-1, None)].update({ft_row_style: xl_cell_style_simple})
            self.row_formatter[(-1, None)].update({ft_row_font: None})

        """ Наконец, назначим типы столбцов, если была указана таблица-шаблон """
        if self.template_was_used:
            self.set_types(**self.guess_col_types(template_data[template_row], keys_ready=True))

        self.sync_segments()

    def __str__(self) -> str:  # Formatter
        about = ''
        about, _l = add_str(about, '{:<17}'.format('Column Formatter:'))
        about, _l = add_str(about, '{:<17}'.format('-' * _l))

        for col_key, col_dic in self.col_formatter.items():
            about, _l = add_str(about, ' {}:'.format(quote(col_key)))
            for col_fmt_key, col_fmt_value in col_dic.items():
                about, _l = add_str(about, '{:<6}{:<7}: {}'.format(' ', quote(col_fmt_key), quote(col_fmt_value)),
                                    strips=False)

        about, _l = add_str(about, '')
        about, _l = add_str(about, '{:<14}'.format('Row Formatter:'))
        about, _l = add_str(about, '{:<14}'.format('-' * _l))

        for row_key, row_dic in self.row_formatter.items():
            about, _l = add_str(about, ' ({}, {}):'.format(quote(row_key[0]), quote(row_key[1])))
            for row_fmt_key, row_fmt_value in row_dic.items():
                about, _l = add_str(about, '{:<12}{:<6}: {}'.format(' ', quote(row_fmt_key), quote(row_fmt_value)),
                                    strips=False)
        return about

    def __add__(self, other):  # Formatter
        new = Formatter()

        """ Форматтер столбцов всегда должен обновляться первым. И только потом - форматтер строк. 
        Иначе будет конфликт ключей столбцов"""

        _formatter = self.get_col_formatter()
        new.set_col_formatter(_formatter, join=True)

        _formatter = self.get_row_formatter()
        new.set_row_formatter(_formatter)

        _formatter = other.get_col_formatter()
        new.set_col_formatter(_formatter, join=True)

        _formatter = other.get_row_formatter()
        new.set_row_formatter(_formatter)
        return new

    def guess_col_types(self, data_row: list | tuple | dict, keys_ready=False) -> dict | list:  # Formatter
        """
        Тестирует Python-типы значений строки и возвращает словарь Excel-типов столбцов {key1:type1,...}

        Parameters
        ----------
        data_row : строка в формате списка или словаря (!!! предполагается, что проверки уже пройдены и на входе только
                   или список или словарь)
        Returns
        -------
        словарь типов столбцов {key1:type1,...}
        """

        trans = {
            int: xl_cell_type_int,
            float: xl_cell_type_float,
            str: xl_cell_type_text
        }
        """ !!! предполагается, что проверки уже пройдены и на входе только или список/кортеж или словарь """
        if isinstance(data_row, list) or isinstance(data_row, tuple):
            _values_list = list(data_row)
        else:
            _values_list = list(data_row.values())

        _types_list = []
        _type_dict = dict()

        for value in _values_list:
            _types_list.append(trans.get(type(value), xl_cell_type_gen))

        if keys_ready:
            _keys_list = self.get_keys_list()
            if len(_types_list) != len(_keys_list):
                raise TableError(f'{class_name(self)}: Размер строки не совпадает с числом ключей\n'
                                 f'Строка: {_values_list}\n'
                                 f'Ключи: {_keys_list}\n'
                                 f'{inspect_info()}')
            else:
                return dict(zip(_keys_list, _types_list))
        else:
            return _types_list

    """---------------------------------------------------------------------------------------------------------------- 
    *** CACHE METHODS >>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>> 
    ----------------------------------------------------------------------------------------------------------------"""

    def sync_segments(self):  # Formatter
        """ Предназначен для создания кэша стилей и типов ячеек, что позволит ускорить форматирование листов.
            Но придется следить за тем, что бы кэш обновлялся при внесении изменений в форматтер при вызове его методов
        Returns
        -------
        None или exception
        """

        """ Сначала - установим кещ типов столбцов """
        self.cached_types_list = self.get_types_list()  # типы столбцов запомнили

        _len = len(self.cached_types_list)  # число столбцов
        _dup = list()
        _row_cache = {}  # заготовка нового словаря кеша

        for seg_tup in self.row_formatter.keys():
            row = seg_tup[0]
            if row not in _dup:
                _dup.append(row)
                _row_styles_lst = []
                for col in range(_len):
                    _row_styles_lst.append(self.get_row_cell_style(row, col))
                _row_cache.update({row: _row_styles_lst})
        self.cached_styles_dic = _row_cache

    def sync_indexes(self):  # Formatter
        _ix = {}

        """ обновляем константу self.columns """
        self.columns = len(self.col_formatter)

        if self.columns > 0:
            """ Перестраиваем индекс  """
            _index = 0
            for key in self.col_formatter.keys():
                _ix.update({_index: key})
                _ix.update({key: _index})
                _index += 1

            self.ix = _ix

            """ добавляем индекс -1. Этот индекс -1 изначально нужен для случая пустого форматтера, когда словарь  
                индексов =   {-1:None, None: -1}. Выражение self.ix[0], часто использовалось в методах форматтера
                строк для обращения к ключу первого столбца, но было несовместимо со случаем пустого (только что 
                инициализированного без аргументов форматтера. Поэтому self.ix[0] был заменен на self.ix[-1] и теперь 
                это выражение совместимо как с пустым, так и заполненным форматтером строк. В случае  заполненного 
                форматтера ссылка на -1 эквивалентна ссылке на 0 """

            self.ix.update({-1: self.ix[0]})

            """ форматтер мог быть изначально создан без аргументов, то есть у него при создании не было ключей 
                столбцов, а значит у него не было ключа столбца в формате строки по умолчанию. Вместо него был None 
                (см. __init__).  Что бы избежать сообщений об ошибках - находим формат строки по умолчанию и 
                пересоздаем его с правильным ключом строки ключа столбца """
            if (-1, None) in self.row_formatter:
                _ft_row_dic = self.row_formatter.pop((-1, None))
                self.row_formatter.update({(-1, self.ix[0]): _ft_row_dic})

        """ И обновим кеш типов столбцов """
        self.cached_types_list = self.get_types_list()  # типы столбцов запомнили

    def ctype(self, col_index: int) -> any:  # Formatter
        """
        Получает тип из кеша типов
        Parameters
        ----------
        col_index : номер столбца (от 0)

        Returns
        -------
        тип столбца
        """
        try:
            return self.cached_types_list[col_index]
        except IndexError:
            return excel_default_type

    def cstyle(self, row_index: int, col_index: int) -> any:  # Formatter
        """
        Получает стиль ячейки из кеша стилей
        Parameters
        ----------
        row_index :
        col_index :

        Returns
        -------
        стиль ячейки

        """
        if row_index not in self.cached_styles_dic:
            row_index = -1
        try:
            return self.cached_styles_dic[row_index][col_index]
        except IndexError:
            return excel_default_style

    def get_keys_list(self) -> list:  # Formatter
        ret = []
        for key in self.col_formatter.keys():
            ret.append(str(key))
        return ret

    """---------------------------------------------------------------------------------------------------------------- 
    *** COLUMNS METHODS>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>> 
    ----------------------------------------------------------------------------------------------------------------"""

    def get_columns_fte_list(self, col_fte_key):  # Formatter
        """
        Возвращает спикок заданного элемента формата для всех столбцов - от первого(0) до последнего.
        fte - это format element (заголовок, тип, значение по умолчанию, функция)

        Parameters
        ----------
        col_fte_key :ключ словаря форматов, определяющий элемент формата ( ft_col_title, ft_col_type, ft_col_default,
                                                                        ft_col_func)

        Returns
        -------
        список элементов формата от первого столбца(0/[A]) до последнего
        """
        ret = []
        for dic in self.col_formatter.values():
            ret.append(dic[col_fte_key])
        return ret

    def get_columns_fte_dict(self, col_fte_key):
        """
        Возвращает словарь заданного элемента формата для всех столбцов - от первого(0) до последнего
        fte - это format element (заголовок, тип, значение по умолчанию, функция)

        Parameters
        ----------
        col_fte_key :ключ словаря форматов, определяющий элемент формата ( ft_col_title, ft_col_type, ft_col_default,
                                                                        ft_col_func)

        Returns
        -------
        словарь элементов формата от первого столбца(0/[A]) до последнего {<ключ_столбца0>:<элемент формата>,...,
                                                                          <ключ_столбцаN>:<элемент формата>}
        """
        ret = {}
        for key, dic in self.col_formatter.items():
            ret.update({key: dic[col_fte_key]})
        return duplicate(ret)

    def set_columns_ftes(self, col_fte_key: str, col_fte_dic: dict) -> None:  # Formatter
        """
        Устанавливает элементы формата col_fte_key согласно словарю {
                                                                    <ключ столбца>:<значение fte столбца>,..,
                                                                    <ключ столбцаN>:<значение fte столбцаN>
                                                                    }
        Parameters
        ----------
        col_fte_dic : словарь элемента форматов для всех столбцов {<ключ столбца>, <значение данного типа fte>}
        col_fte_key : ключ элемента формата ( ( ft_col_title, ft_col_type, ft_col_default, ft_col_func)

        Returns
        -------
        None
        """

        self.validate_col_keys(col_fte_dic)
        for key, value in col_fte_dic.items():
            self.col_formatter[key][col_fte_key] = value

        self.sync_indexes()  # !

    """ 1) titles """

    def set_titles(self, **titles):  # Formatter
        self.set_columns_ftes(ft_col_title, titles)

    def get_titles_list(self):  # Formatter
        return self.get_columns_fte_list(ft_col_title)

    def get_titles_dict(self):  # Formatter
        return self.get_columns_fte_dict(ft_col_title)

    """ 2) types """

    def set_types(self, **types):  # Formatter
        self.set_columns_ftes(ft_col_type, types)

    def get_types_list(self):  # Formatter
        return self.get_columns_fte_list(ft_col_type)

    def get_types_dict(self):  # Formatter
        return self.get_columns_fte_dict(ft_col_type)

    """ 3) defaults """

    def set_defaults(self, **defaults):  # Formatter
        self.set_columns_ftes(ft_col_default, defaults)

    def get_defaults_list(self):  # Formatter
        return self.get_columns_fte_list(ft_col_default)

    def get_defaults_dict(self):  # Formatter
        return self.get_columns_fte_dict(ft_col_default)

    """ 4) funcs """

    def set_funcs(self, **funcs):  # Formatter
        self.set_columns_ftes(ft_col_func, funcs)

    def get_funcs_list(self):  # Formatter
        return self.get_columns_fte_list(ft_col_func)

    def get_funcs_dict(self):  # Formatter
        return self.get_columns_fte_dict(ft_col_func)

    """ ------------------------------------------------------------------------------------------------------------"""

    def set_col_formatter(self, _col_formatter: dict, join=False) -> None:  # Formatter
        """
        Устанавливает или обновляет (join=true) словарь форматтера столбцов
        Parameters
        ----------
        _col_formatter : входной форматтер столбцов на входе
        join : если True  - то обновляет существующие форматы столбцов, или добавляет форматы столбцов, если их не было
               если False - то обновляет существующие форматы столбцов. Если встретился столбец (ключ), которого нет
                            в форматтере - генерирует исключение TableError

        Returns
        -------
        None
        """

        _in, _out = self.match_col_keys(_col_formatter)

        if join and (len(_in) > 0):
            raise TableError(f'{class_name(self)}: При слиянии форматов обнаружено совпадение ключей с существующими: '
                             f'{", ".join(_in)}'
                             f'\n{inspect_info()}')
        elif (not join) and (len(_out) > 0):
            raise TableError(f'{class_name(self)}: Ключ(и) {", ".join(_out)} - не найден(ы). Доступны ключи:'
                             f' {" ,".join([str(k) for k in self.col_formatter.keys()])}'
                             f'\n{inspect_info()}')

        for _col_key, _col_fte_dic in _col_formatter.items():
            if _col_key not in self.col_formatter:
                self.col_formatter.update({_col_key: {}})
            for _col_fte_key, _col_fte_value in _col_fte_dic.items():
                self.col_formatter[_col_key].update({_col_fte_key: _col_fte_value})

        self.sync_indexes()

    def get_col_formatter(self) -> dict:  # Formatter
        """
        Возвращает словарь форматтера столбцов
        Returns
        -------
        словарь форматтера столбцов
        """
        _col_formatter = {}
        for col_key, col_dic in self.col_formatter.items():
            _col_formatter.update({col_key: {}})
            for col_fmt_key, col_fmt_value in col_dic.items():
                _col_formatter[col_key].update({col_fmt_key: col_fmt_value})
        return duplicate(_col_formatter)

    """----------------------------------------------------------------------------------------------------------------
    *** VERIFY METHODS >>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>
    ----------------------------------------------------------------------------------------------------------------"""

    def match_col_keys(self, dic) -> tuple[list, list]:  # Formatter
        keys_in = []
        keys_out = []
        if self.columns > 0:
            for key in dic.keys():
                if key in self.col_formatter:
                    keys_in.append(key)
                else:
                    keys_out.append(key)
        return keys_in, keys_out

    def validate_col_keys(self, dic):  # Formatter
        _in, _out = self.match_col_keys(dic)
        if len(_out) > 0:
            raise TableError(f'{class_name(self)}: Ключ(и) {", ".join(_out)} - не найден(ы). Доступны ключи:'
                             f' {" ,".join([str(k) for k in self.col_formatter.keys()])}'
                             f'\n{inspect_info()}')

    """---------------------------------------------------------------------------------------------------------------- 
    *** ROW METHODS >>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>> 
    ----------------------------------------------------------------------------------------------------------------"""

    def set_row_fte_segment(self, row_index: int, col_key: any, row_fte_key: str, row_fte_value: any):  # Formatter

        """ Сначала проверяем ключ столбца (искусственно создаем словарь, где важен только ключ.
                                            Єто что бы воспользоваться уже готовой validate_col_keys) """

        self.validate_col_keys({col_key: None})  # если col_key не найден - будет исключение TableError

        """ Теперь проверяем, не является ли fte уже таким, как устанавливается (избегаем повторов одинаковых наборов 
            fte в смежных сегментах """

        """ получаем текущий словарь формата сегмента """
        seg_fmt_dic = self.get_row_cell_ft(row_index, self.ix[col_key])
        curr_fte_value = seg_fmt_dic[row_fte_key]

        if curr_fte_value != row_fte_value:
            """ формируем кортеж ключа сегмента """
            seg_key_tup = (row_index, col_key)
            """ устанавливаем новое значение заданного fte """
            seg_fmt_dic[row_fte_key] = row_fte_value
            """ вносим его в форматтер """
            self.row_formatter.update({seg_key_tup: seg_fmt_dic})

        self.sync_segments()  # !

    def set_row_style_segment(self, row_index: int, col_key: any, row_style: str):  # Formatter
        self.set_row_fte_segment(row_index, col_key, ft_row_style, row_style)

    def set_row_style(self, row_index: int, row_style: str):  # Formatter
        self.set_row_style_segment(row_index, self.ix[0], row_style)

    def get_row_segments_list(self, row_index: int) -> list[tuple]:  # Formatter
        """
            Этот метод возвращает список кортежей (col_index, row_ft_dic) отсортированный по номерам
            столбцов (col_index) от большего к меньшему. Формат кортежа: (номер первого столбца сегмента, словарь fte)

        Parameters
        ----------
        row_index : номер строки (0-...)

        Returns
        -------
        Список кортежей (col_index, row_ft_dic) отсортированный по номерам столбцов (col_index) от большего к меньшему.
        Формат кортежа: (<номер первого столбца сегмента>, <словарь fte>)
        """

        _row_segments_ft_list: list = []

        """ Сначала проверяем, есть ли сегмент этой строки, который начинается с первого столбца (col_index = 0).
            Если есть - заносим его в список. Если нет - заносим в список номер первого столбца со стилем строки по 
            умолчанию (-1).
        """
        if (row_index, self.ix[-1]) not in self.row_formatter:
            _seg_ft_dic = self.row_formatter[(-1, self.ix[-1])].copy()
            _row_segments_ft_list.append((0, _seg_ft_dic))

        """ Теперь ищем остальные сегменты строки """
        for _seg_key_tup, _seg_ft_dic in self.row_formatter.items():
            if _seg_key_tup[0] == row_index:
                _col_index = self.ix[_seg_key_tup[1]]
                _row_segments_ft_list.append((_col_index, _seg_ft_dic))
        # и сортируем по убыванию (что бы проще потом искать попадание столбца в сегмент)
        _row_segments_ft_list.sort(key=lambda _seg: _seg[0], reverse=True)
        return duplicate(_row_segments_ft_list)

    def get_row_cell_ft(self, row: int, col: int) -> any:  # Formatter
        """
        Возвращает row ft ячейки
        Parameters
        ----------
        row : номер строки (от 0 )
        col : номер столбца (от 0)
        Returns
        -------
        fte
        """
        seg_tup_list = self.get_row_segments_list(row)  # [(<col_index>, {'style': 'Output', 'text': None}),...]

        if len(seg_tup_list) == 1:  # не столько оптимизация, сколько обработка случая, когда форматтер колонок пуст,
            # потому что в этом случае номер столбца после трансляции из ключа None будет -1
            # и сортировка (см.ниже) не получится и тогда: "TableError(f'Ошибка поиска словаря..
            row_cell_ft = seg_tup_list[0][1]
        else:
            for _seg_tup in seg_tup_list:
                if _seg_tup[0] <= col:
                    row_cell_ft = _seg_tup[1]
                    break
            else:
                raise TableError(
                    f'{class_name(self)}: Ошибка поиска словаря форматов строки [{row}] для столбца {col}.\n'
                    f'{inspect_info()}')
        return duplicate(row_cell_ft)

    def get_row_cell_fte(self, row: int, col: int, row_fte_key: str) -> any:  # Formatter
        return self.get_row_cell_ft(row, col)[row_fte_key]

    def get_row_cell_style(self, row: int, col: int) -> str:  # Formatter
        """ Возвращает стиль ячейки """
        return self.get_row_cell_fte(row, col, ft_row_style)

    """ ------------------------------------------------------------------------------------------------------------"""

    def get_row_formatter(self) -> dict:  # Formatter
        """
        Возврашщает форматтер строк

        Returns
        -------
        словарь форматтера строк

        """
        _row_formatter = {}
        for row_key, row_dic in self.row_formatter.items():  # пересобираем, иначе прелести mutable или делать deep copy
            _row_formatter.update({row_key: {}})
            for row_fte_key, row_fte_value in row_dic.items():
                _row_formatter[row_key].update({row_fte_key: row_fte_value})
        return duplicate(_row_formatter)

    def set_row_formatter(self, _row_formatter: dict) -> None:  # Formatter
        """
        Устанавливает или обновляет словарь форматтера строк
        Parameters
        ----------
        _row_formatter : входной форматтер строк

        Returns
        -------
        None
        """

        """ Сначала проверяем наличие ключей столбцов. для этого сразу построим проверочный словарь в формате 
            validate_col_keys(). """
        _tst_dic = {}
        for _row_key_tup in _row_formatter.keys():
            _tst_dic.update({_row_key_tup[1]: None})
        self.validate_col_keys(_tst_dic)  # если какой-то из ключей не найден - будет exception

        for _row_key_tup, _row_ft_dic in _row_formatter.items():
            """ проверим, во избежание излишнего сегментирования, не является ли очередной словарь ft равным 
                эквивалентному (действующему) словарю сегмента """
            _curr_ft_dic = self.get_row_cell_ft(_row_key_tup[0], self.ix[_row_key_tup[1]])
            if _row_ft_dic != _curr_ft_dic:
                if _row_key_tup not in self.row_formatter:
                    self.row_formatter.update({_row_key_tup: {}})
                for _row_fte_key, _row_fte_value in _row_ft_dic.items():
                    self.row_formatter[_row_key_tup].update({_row_fte_key: _row_fte_value})

        self.sync_segments()  # !

    def format_sheet(self, worksheet: openpyxl.worksheet.worksheet.Worksheet, table: list,  # Formatter
                     add_keys=False, add_titles=False) -> None:
        """
        С помощью этого метода, в случае необходимости, можно отформатировать лист excel и заполнить его данными.
        из таблицы (list of list/list of dict)

        Parameters
        ----------
        worksheet:  лист openpyxl
        table:      list of list - таблица данных для листа
        add_keys:   Если True - добавить ключи в заголовок таблицы лист Excel
        add_titles: Если True - добавить заголовки в (под)заголовок таблицы лист Excel.
                (*) Если указан и add_keys = True и add_titles = True, то первой строкой таблицы будут ключи,
                    а второй - заголовки
        Returns
        -------
        """

        ws_table = list()
        self.rows = 0

        """ В качестве принимаемой таблицы может быть список списков (список строк, где строки - списки) или 
            список словарей. Если размер строки списка списков меньше числа столбцов форматтера - будет исключение, а 
            если больше - то строка будет усечена до размера словаря столбцов форматтера.

            В случае, если таблица задана списком словарей, то из каждого словаря строк будут выбраны  значения ключей
            форматтера столбцов в порядке следования ключей в форматтере. Если ключ не не найден - будет инициировано
            исключение. """

        for data in table:
            if isinstance(data, list):
                data_row = data

            elif isinstance(data, dict):
                data_row = []
                for key in self.col_formatter.keys():
                    if key in data_row:
                        data_row.append(data_row[key])
            else:
                raise TableError(f'{class_name(self)}: Форматтер не поддерживает тип представления {type(data)} '
                                 f'принимаемой строки.\n{inspect_info()}')
            if len(data_row) > 0:
                # ws_table.append(data_row[0: self.columns])
                ws_table.append(data_row)
                self.rows += 1

        if add_titles:
            ws_table.insert(0, self.get_titles_list())
            self.rows += 1
        if add_keys:
            ws_table.insert(0, self.get_keys_list())
            self.rows += 1

        if self.rows == 0:
            raise TableError(f'{class_name(self)}: Форматтер получил пустую таблицу.\n'
                             f'{inspect_info()}.\n')
        else:
            row_index = 0
            while row_index < self.rows:
                ws_row = ws_table[row_index]
                ws_row_len = len(ws_row)

                formatted_row = []
                col_index = 0
                while col_index < ws_row_len:
                    cell = WriteOnlyCell(worksheet, value=ws_row[col_index])
                    if col_index < self.columns:
                        cell.style = self.cstyle(row_index, col_index)
                        cell.number_format = self.ctype(col_index)
                    #  column = worksheet.column_dimensions[get_column_letter(col_index + 1)]
                    #  column.number_format = self.ctype(col_index)
                    formatted_row.append(cell)
                    col_index += 1

                worksheet.append(formatted_row)
                row_index += 1
